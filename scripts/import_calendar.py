import csv
import hashlib
import io
import json
import os
import re
from dataclasses import asdict, dataclass
from datetime import date, datetime, time, timedelta
from pathlib import Path
from typing import Optional
from urllib.parse import quote

import requests
from bs4 import BeautifulSoup

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

from config import (
    CLUB_ID,
    HANDBALLNET_CLUB_SLUG,
    CLUB_NAME,
    TIMEZONE,
    DATE_FROM,
    DATE_TO,
    HALLS,
    DEFAULT_GAME_DURATION_MINUTES,
    TRAINING_CSV,
    WEEKEND_XLSX,
    WEEKEND_HALL_MAP,
    EXTRA_EVENTS_CSV_URL,
    EXTRA_HALL_MAP,
    EXTRA_TYPE_MAP,
)


@dataclass
class CalendarEvent:
    id: str
    title: str
    start: str
    end: str
    hall_id: str
    hall: str
    type: str
    source: str
    location: str = ""
    description: str = ""
    url: str = ""
    color: str = ""


def clean_text(value: str) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def make_id(prefix: str, *parts: object) -> str:
    raw = "|".join(str(p) for p in parts)
    digest = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]
    return f"{prefix}-{digest}"


def to_iso(dt: datetime) -> str:
    return dt.strftime("%Y-%m-%dT%H:%M:%S")


def parse_date(value: str) -> Optional[date]:
    value = clean_text(value)

    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            pass

    return None


def parse_time(value: str) -> Optional[time]:
    value = clean_text(value).replace(" Uhr", "").replace("Uhr", "").strip()
    value = value.replace(".", ":")

    match = re.search(r"(\d{1,2})[:](\d{2})", value)
    if not match:
        return None

    hour = int(match.group(1))
    minute = int(match.group(2))

    if hour < 0 or hour > 23 or minute < 0 or minute > 59:
        return None

    return time(hour, minute)


def parse_game_start(text: str) -> Optional[datetime]:
    text = clean_text(text)

    patterns = [
        r"Spielbeginn\s+([A-Za-zÄÖÜäöüß]+,\s*)?(\d{1,2}\.\d{1,2}\.\d{4})\s*-\s*(\d{1,2}:\d{2})",
        r"Spielbeginn\s+([A-Za-zÄÖÜäöüß]+,\s*)?(\d{1,2}\.\d{1,2}\.\d{2})\s*-\s*(\d{1,2}:\d{2})",
    ]

    for pattern in patterns:
        match = re.search(pattern, text, flags=re.I)
        if not match:
            continue

        date_value = parse_date(match.group(2))
        time_value = parse_time(match.group(3))

        if date_value and time_value:
            return datetime.combine(date_value, time_value)

    return None


def extract_game_number(text: str, hall_id: str, start: datetime) -> str:
    text = clean_text(text)

    match = re.search(r"Spielnummer\s+([A-Za-z0-9_-]+)", text, flags=re.I)
    if match:
        return match.group(1).strip()

    match = re.search(r"Spiel-Nr\.?\s*([A-Za-z0-9_-]+)", text, flags=re.I)
    if match:
        return match.group(1).strip()

    return make_id("game", hall_id, start.isoformat(), text)


def extract_between(text: str, start_label: str, end_labels: list[str]) -> str:
    text = clean_text(text)

    start_match = re.search(re.escape(start_label), text, flags=re.I)
    if not start_match:
        return ""

    start_pos = start_match.end()
    end_pos = len(text)

    for label in end_labels:
        end_match = re.search(re.escape(label), text[start_pos:], flags=re.I)
        if end_match:
            end_pos = min(end_pos, start_pos + end_match.start())

    return clean_text(text[start_pos:end_pos])


def extract_game_title(text: str, start: datetime) -> str:
    text = clean_text(text)

    staffel = extract_between(
        text,
        "Staffel / Runde",
        ["Spielnummer", "Spielbeginn", "Halle", "Heim", "Gast"],
    )

    teams = []

    for label in ["Heim", "Gast", "Heimmannschaft", "Gastmannschaft"]:
        value = extract_between(
            text,
            label,
            ["Heim", "Gast", "Heimmannschaft", "Gastmannschaft", "Spielnummer", "Spielbeginn", "Halle", "Staffel / Runde"],
        )
        if value and len(value) < 80:
            teams.append(value)

    teams = [t for t in teams if t and t.lower() not in ["heim", "gast"]]

    if len(teams) >= 2:
        team_title = f"{teams[0]} - {teams[1]}"
    elif len(teams) == 1:
        team_title = teams[0]
    else:
        team_title = ""

    if staffel and team_title:
        return f"{staffel} · {team_title}"

    if team_title:
        return team_title

    if staffel:
        return staffel

    return f"Spiel {start.strftime('%d.%m.%Y %H:%M')}"


def fetch_handballnet_games() -> list[CalendarEvent]:
    """Fetch handball.net games in weekly chunks from old and new URL structures.

    Alte Struktur:
    https://www.handball.net/vereine/<CLUB_ID>/spielplan

    Neue Struktur:
    https://handball.net/club/<HANDBALLNET_CLUB_SLUG>

    Doppelte Spiele werden über die Spielnummer/Event-ID entfernt.
    """

    headers = {
        "User-Agent": "Mozilla/5.0 hallenkalender-import/1.0 (+https://github.com/)",
        "Accept-Language": "de-DE,de;q=0.9,en;q=0.8",
    }

    events: list[CalendarEvent] = []
    seen: set[str] = set()

    max_pages = 20
    current = DATE_FROM

    while current <= DATE_TO:
        chunk_to = min(current + timedelta(days=6), DATE_TO)

        candidate_base_urls = [
            (
                "alte handball.net URL",
                f"https://www.handball.net/vereine/{CLUB_ID}/spielplan"
                f"?dateFrom={current.isoformat()}&dateTo={chunk_to.isoformat()}",
            ),
            (
                "neue handball.net Club URL",
                f"https://handball.net/club/{HANDBALLNET_CLUB_SLUG}"
                f"?dateFrom={current.isoformat()}&dateTo={chunk_to.isoformat()}",
            ),
        ]

        print(f"handball.net Zeitraum: {current.isoformat()} bis {chunk_to.isoformat()}")

        for source_name, base_url in candidate_base_urls:
            print(f"Teste {source_name}: {base_url}")

            pages_without_new_events = 0

            for page in range(1, max_pages + 1):
                if page == 1:
                    url = base_url
                else:
                    url = f"{base_url}&page={page}"

                print(f"Fetching handball.net page {page}: {url}")

                try:
                    response = requests.get(url, headers=headers, timeout=30)
                    response.raise_for_status()
                except Exception as exc:
                    print(f"WARNING: handball.net request failed for {url}: {exc}")
                    break

                soup = BeautifulSoup(response.text, "html.parser")
                text_blocks: list[str] = []

                for tag in soup.find_all(["article", "section", "li", "tr", "div"]):
                    txt = clean_text(tag.get_text(" "))

                    if len(txt) < 80 or len(txt) > 2500:
                        continue

                    if "Spielbeginn" not in txt:
                        continue

                    if "Spielnummer" not in txt:
                        continue

                    if not any(hall_id in txt for hall_id in HALLS):
                        continue

                    text_blocks.append(txt)

                before_count = len(events)

                for block in text_blocks:
                    block_without_update = re.split(r"letztes\s+Update", block, flags=re.I)[0]

                    for hall_id, hall in HALLS.items():
                        if hall_id not in block_without_update:
                            continue

                        start = parse_game_start(block_without_update)

                        if not start:
                            continue

                        if start.date() < DATE_FROM or start.date() > DATE_TO:
                            continue

                        end = start + timedelta(minutes=DEFAULT_GAME_DURATION_MINUTES)

                        game_no = extract_game_number(block_without_update, hall_id, start)
                        event_id = f"handballnet-{game_no}"

                        if event_id in seen:
                            continue

                        seen.add(event_id)

                        title = extract_game_title(block_without_update, start)

                        events.append(
                            CalendarEvent(
                                id=event_id,
                                title=title,
                                start=to_iso(start),
                                end=to_iso(end),
                                hall_id=hall_id,
                                hall=hall["name"],
                                type="game",
                                source="handball.net",
                                location=hall["name"],
                                description=(
                                    f"Quelle: handball.net | {CLUB_NAME} | "
                                    f"Hallennummer {hall_id} | Spielnummer {game_no}"
                                ),
                                url=url,
                                color=hall.get("color", ""),
                            )
                        )

                new_events = len(events) - before_count

                print(
                    f"{source_name}, Zeitraum {current.isoformat()} bis {chunk_to.isoformat()}, "
                    f"page {page}: {new_events} new hall games"
                )

                if new_events == 0:
                    pages_without_new_events += 1
                else:
                    pages_without_new_events = 0

                if pages_without_new_events >= 4:
                    print(
                        f"Stopping handball.net pagination for {source_name} "
                        "after 4 pages without new hall games in this week."
                    )
                    break

        current = chunk_to + timedelta(days=1)

    print(f"handball.net games gesamt nach Wochenimport: {len(events)}")

    return sorted(events, key=lambda e: e.start)


def weekday_to_int(value: str) -> Optional[int]:
    value = clean_text(value).upper()

    mapping = {
        "MO": 0,
        "MON": 0,
        "MONTAG": 0,
        "DI": 1,
        "TU": 1,
        "TUE": 1,
        "DIENSTAG": 1,
        "MI": 2,
        "WE": 2,
        "WED": 2,
        "MITTWOCH": 2,
        "DO": 3,
        "TH": 3,
        "THU": 3,
        "DONNERSTAG": 3,
        "FR": 4,
        "FRI": 4,
        "FREITAG": 4,
        "SA": 5,
        "SAT": 5,
        "SAMSTAG": 5,
        "SO": 6,
        "SU": 6,
        "SUN": 6,
        "SONNTAG": 6,
    }

    return mapping.get(value)


def load_training_events() -> list[CalendarEvent]:
    path = Path(TRAINING_CSV)

    if not path.exists():
        print(f"training csv not found: {TRAINING_CSV}")
        return []

    events: list[CalendarEvent] = []

    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)

        for row_no, row in enumerate(reader, start=2):
            event_type = clean_text(row.get("type", "")) or "training"
            team = clean_text(row.get("team", "")) or "Training"
            hall_id = clean_text(row.get("hall_id", ""))
            weekday_raw = clean_text(row.get("weekday", ""))
            start_raw = clean_text(row.get("start_time", ""))
            end_raw = clean_text(row.get("end_time", ""))
            date_from_raw = clean_text(row.get("date_from", ""))
            date_to_raw = clean_text(row.get("date_to", ""))
            notes = clean_text(row.get("notes", ""))

            if hall_id not in HALLS:
                print(f"WARNING: training row {row_no} skipped, unknown hall_id: {hall_id}")
                continue

            weekday = weekday_to_int(weekday_raw)
            start_time = parse_time(start_raw)
            end_time = parse_time(end_raw)
            row_date_from = parse_date(date_from_raw)
            row_date_to = parse_date(date_to_raw)

            if weekday is None or not start_time or not end_time or not row_date_from or not row_date_to:
                print(f"WARNING: training row {row_no} skipped, invalid date/time/weekday")
                continue

            valid_from = max(row_date_from, DATE_FROM)
            valid_to = min(row_date_to, DATE_TO)

            if valid_from > valid_to:
                continue

            current = valid_from

            while current.weekday() != weekday:
                current += timedelta(days=1)

            hall = HALLS[hall_id]

            while current <= valid_to:
                start_dt = datetime.combine(current, start_time)
                end_dt = datetime.combine(current, end_time)

                if end_dt <= start_dt:
                    end_dt += timedelta(days=1)

                if event_type == "training":
                    title = f"Training {team}"
                elif event_type == "blocked":
                    title = f"Belegt: {team}"
                else:
                    title = team

                events.append(
                    CalendarEvent(
                        id=make_id("training", event_type, team, hall_id, current, start_time, end_time),
                        title=title,
                        start=to_iso(start_dt),
                        end=to_iso(end_dt),
                        hall_id=hall_id,
                        hall=hall["name"],
                        type=event_type,
                        source="trainings.csv",
                        location=hall["name"],
                        description=notes,
                        color=hall.get("color", ""),
                    )
                )

                current += timedelta(days=7)

    print(f"training events: {len(events)}")
    return events


def excel_cell_text(value: object) -> str:
    if value is None:
        return ""
    return clean_text(str(value))


def parse_excel_date(value: object) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = excel_cell_text(value)

    match = re.search(r"(\d{1,2}\.\d{1,2}\.\d{2,4})", text)
    if match:
        return parse_date(match.group(1))

    return parse_date(text)


def parse_excel_time_range(value: object) -> tuple[Optional[time], Optional[time]]:
    text = excel_cell_text(value)

    matches = re.findall(r"(\d{1,2})[:.](\d{2})", text)
    if len(matches) < 2:
        return None, None

    start = time(int(matches[0][0]), int(matches[0][1]))
    end = time(int(matches[1][0]), int(matches[1][1]))

    return start, end


def load_weekend_excel_events() -> list[CalendarEvent]:
    if not WEEKEND_XLSX:
        print("weekend excel disabled")
        return []

    path = Path(WEEKEND_XLSX)

    if not path.exists():
        print(f"weekend excel not found: {WEEKEND_XLSX}")
        return []

    if load_workbook is None:
        print("WARNING: openpyxl not installed, weekend excel skipped")
        return []

    workbook = load_workbook(path, data_only=True)
    events: list[CalendarEvent] = []

    for sheet in workbook.worksheets:
        current_date: Optional[date] = None
        hall_columns: dict[int, str] = {}

        for row in sheet.iter_rows():
            values = [cell.value for cell in row]

            first_value = values[0] if values else None
            possible_date = parse_excel_date(first_value)

            if possible_date:
                current_date = possible_date

            for index, value in enumerate(values):
                text = excel_cell_text(value)

                if text in WEEKEND_HALL_MAP:
                    hall_columns[index] = WEEKEND_HALL_MAP[text]

            if not current_date:
                continue

            for col_index, hall_id in hall_columns.items():
                if hall_id not in HALLS:
                    continue

                time_col_index = max(0, col_index - 2)
                title_col_index = col_index

                time_value = values[time_col_index] if time_col_index < len(values) else None
                title_value = values[title_col_index] if title_col_index < len(values) else None

                title = excel_cell_text(title_value)

                if not title:
                    continue

                if title in WEEKEND_HALL_MAP:
                    continue

                if len(title) < 2:
                    continue

                start_time, end_time = parse_excel_time_range(time_value)

                if not start_time or not end_time:
                    continue

                if current_date < DATE_FROM or current_date > DATE_TO:
                    continue

                start_dt = datetime.combine(current_date, start_time)
                end_dt = datetime.combine(current_date, end_time)

                if end_dt <= start_dt:
                    end_dt += timedelta(days=1)

                hall = HALLS[hall_id]

                events.append(
                    CalendarEvent(
                        id=make_id("weekend", sheet.title, current_date, hall_id, start_time, end_time, title),
                        title=title,
                        start=to_iso(start_dt),
                        end=to_iso(end_dt),
                        hall_id=hall_id,
                        hall=hall["name"],
                        type="weekend",
                        source="weekend excel",
                        location=hall["name"],
                        description=f"Quelle: {WEEKEND_XLSX}",
                        color=hall.get("color", ""),
                    )
                )

    print(f"weekend excel events: {len(events)}")
    return events


def get_row_value(row: dict[str, str], names: list[str]) -> str:
    normalized = {clean_text(k).lower(): v for k, v in row.items()}

    for name in names:
        key = clean_text(name).lower()
        if key in normalized:
            return clean_text(normalized[key])

    return ""


def load_extra_events() -> list[CalendarEvent]:
    if not EXTRA_EVENTS_CSV_URL:
        print("extra events: disabled, EXTRA_EVENTS_CSV_URL not configured")
        return []

    print(f"extra events: fetching {EXTRA_EVENTS_CSV_URL}")

    try:
        response = requests.get(EXTRA_EVENTS_CSV_URL, timeout=30)
        response.raise_for_status()
    except Exception as exc:
        print(f"WARNING: extra events fetch failed: {exc}")
        return []

    text = response.text

    if "<html" in text.lower() or "<!doctype html" in text.lower():
        print("WARNING: extra events URL returned HTML instead of CSV")
        return []

    reader = csv.DictReader(io.StringIO(text))
    events: list[CalendarEvent] = []

    for row_no, row in enumerate(reader, start=2):
        title = get_row_value(row, ["Titel", "title", "Name", "Termin"])
        hall_name = get_row_value(row, ["Halle", "hall", "Sporthalle"])
        date_raw = get_row_value(row, ["Datum", "date"])
        start_raw = get_row_value(row, ["Startzeit", "Start", "Beginn"])
        end_raw = get_row_value(row, ["Endzeit", "Ende"])
        type_raw = get_row_value(row, ["Art", "Typ", "type"])
        notes = get_row_value(row, ["Notiz", "Notizen", "Beschreibung", "Info"])

        if not title:
            title = "Zusatztermin"

        hall_id = EXTRA_HALL_MAP.get(hall_name)

        if not hall_id:
            print(f"WARNING: unknown hall in extra event skipped row {row_no}: {hall_name}")
            continue

        if hall_id not in HALLS:
            print(f"WARNING: extra event hall_id not in HALLS skipped row {row_no}: {hall_id}")
            continue

        event_date = parse_date(date_raw)
        start_time = parse_time(start_raw)
        end_time = parse_time(end_raw)

        if not event_date or not start_time or not end_time:
            print(f"WARNING: extra event date/time parse failed row {row_no}")
            continue

        if event_date < DATE_FROM or event_date > DATE_TO:
            continue

        event_type = EXTRA_TYPE_MAP.get(type_raw, "event")

        start_dt = datetime.combine(event_date, start_time)
        end_dt = datetime.combine(event_date, end_time)

        if end_dt <= start_dt:
            end_dt += timedelta(days=1)

        hall = HALLS[hall_id]

        events.append(
            CalendarEvent(
                id=make_id("extra", title, hall_id, event_date, start_time, end_time, notes),
                title=title,
                start=to_iso(start_dt),
                end=to_iso(end_dt),
                hall_id=hall_id,
                hall=hall["name"],
                type=event_type,
                source="Google Form",
                location=hall["name"],
                description=notes,
                color=hall.get("color", ""),
            )
        )

    print(f"extra events: {len(events)}")
    return events


def escape_ics_text(value: str) -> str:
    value = str(value or "")
    value = value.replace("\\", "\\\\")
    value = value.replace("\n", "\\n")
    value = value.replace(",", "\\,")
    value = value.replace(";", "\\;")
    return value


def format_ics_datetime(value: str) -> str:
    dt = datetime.fromisoformat(value)
    return dt.strftime("%Y%m%dT%H%M%S")


def fold_ics_line(line: str) -> list[str]:
    encoded = line.encode("utf-8")

    if len(encoded) <= 75:
        return [line]

    result = []
    current = ""

    for char in line:
        test = current + char

        if len(test.encode("utf-8")) > 75:
            result.append(current)
            current = " " + char
        else:
            current = test

    if current:
        result.append(current)

    return result


def write_ics(path: Path, events: list[CalendarEvent], calendar_name: str) -> None:
    lines = [
        "BEGIN:VCALENDAR",
        "VERSION:2.0",
        "PRODID:-//HSG FONA//Hallenkalender//DE",
        "CALSCALE:GREGORIAN",
        "METHOD:PUBLISH",
        f"X-WR-CALNAME:{escape_ics_text(calendar_name)}",
        f"X-WR-TIMEZONE:{TIMEZONE}",
    ]

    now = datetime.utcnow().strftime("%Y%m%dT%H%M%SZ")

    for event in events:
        uid = f"{event.id}@hsg-fona.github.io"

        lines.extend(
            [
                "BEGIN:VEVENT",
                f"UID:{escape_ics_text(uid)}",
                f"DTSTAMP:{now}",
                f"DTSTART;TZID={TIMEZONE}:{format_ics_datetime(event.start)}",
                f"DTEND;TZID={TIMEZONE}:{format_ics_datetime(event.end)}",
                f"SUMMARY:{escape_ics_text(event.title)}",
                f"LOCATION:{escape_ics_text(event.location or event.hall)}",
                f"DESCRIPTION:{escape_ics_text(event.description)}",
            ]
        )

        if event.url:
            lines.append(f"URL:{escape_ics_text(event.url)}")

        lines.append("END:VEVENT")

    lines.append("END:VCALENDAR")

    folded_lines = []

    for line in lines:
        folded_lines.extend(fold_ics_line(line))

    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\r\n".join(folded_lines) + "\r\n", encoding="utf-8")


def write_json(path: Path, events: list[CalendarEvent]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    data = [asdict(event) for event in events]

    path.write_text(
        json.dumps(data, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def main() -> None:
    print(f"calendar import: {DATE_FROM.isoformat()} bis {DATE_TO.isoformat()}")

    games = fetch_handballnet_games()
    print(f"handball.net games: {len(games)}")

    trainings = load_training_events()
    print(f"training events: {len(trainings)}")

    weekend_events = load_weekend_excel_events()
    print(f"weekend excel events: {len(weekend_events)}")

    extra_events = load_extra_events()
    print(f"extra events: {len(extra_events)}")

    events = games + trainings + weekend_events + extra_events

    events = sorted(events, key=lambda event: event.start)

    print(f"events total: {len(events)}")

    write_json(Path("data/events.json"), events)

    write_ics(Path("calendars/gesamt.ics"), events, "HSG FONA Hallenkalender Gesamt")

    for hall_id, hall in HALLS.items():
        hall_events = [event for event in events if event.hall_id == hall_id]
        slug = hall.get("slug", hall_id)
        path = Path("calendars") / f"{slug}.ics"

        write_ics(path, hall_events, f"HSG FONA {hall['name']}")

        print(f"ics {hall['name']}: {len(hall_events)} events")

    print("calendar import finished")


if __name__ == "__main__":
    main()
